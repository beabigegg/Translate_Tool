"""Term database backed by SQLite."""

from __future__ import annotations

import csv
import json
import logging
import sqlite3
import threading
from datetime import datetime, timezone
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

from app.backend.models.term import Term
from app.backend.config import TERM_INJECT_HIGH_CONFIDENCE_UNVERIFIED, TERM_INJECT_CONF_THRESHOLD

logger = logging.getLogger(__name__)

_DB_LOCK = threading.Lock()

_VALID_STATUSES = {"unverified", "needs_review", "approved", "rejected"}
_INJECTABLE_STATUSES = {"approved"}

_CREATE_TABLE_SQL = """
CREATE TABLE IF NOT EXISTS terms (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_text TEXT NOT NULL,
    target_text TEXT NOT NULL,
    source_lang TEXT NOT NULL,
    target_lang TEXT NOT NULL,
    domain TEXT NOT NULL DEFAULT 'general',
    context_snippet TEXT NOT NULL DEFAULT '',
    confidence REAL NOT NULL DEFAULT 1.0,
    usage_count INTEGER NOT NULL DEFAULT 0,
    status TEXT NOT NULL DEFAULT 'unverified',
    created_at TEXT NOT NULL,
    UNIQUE (source_text, target_lang, domain)
);
"""

_MIGRATE_ADD_STATUS_SQL = """
ALTER TABLE terms ADD COLUMN status TEXT NOT NULL DEFAULT 'unverified';
"""


class TermDB:
    """Persistent term database stored in a local SQLite file."""

    def __init__(self, db_path: Optional[Path] = None) -> None:
        if db_path is None:
            from app.backend.config import DATA_DIR
            db_path = DATA_DIR / "term_db.sqlite"

        self.db_path = Path(db_path)
        self.db_path.parent.mkdir(parents=True, exist_ok=True)
        self._init_db()

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(str(self.db_path), check_same_thread=False)
        conn.row_factory = sqlite3.Row
        return conn

    def _init_db(self) -> None:
        with _DB_LOCK, self._connect() as conn:
            conn.execute(_CREATE_TABLE_SQL)
            conn.commit()
            # Migration: add status column if absent (existing DBs)
            cols = {r[1] for r in conn.execute("PRAGMA table_info(terms)").fetchall()}
            if "status" not in cols:
                conn.execute(_MIGRATE_ADD_STATUS_SQL)
                conn.commit()
                logger.info("[TermDB] Migrated: added status column")

    # ------------------------------------------------------------------
    # Query methods
    # ------------------------------------------------------------------

    def exists(self, source_text: str, target_lang: str, domain: str) -> bool:
        """Return True if (source_text, target_lang, domain) is already stored."""
        with _DB_LOCK, self._connect() as conn:
            row = conn.execute(
                "SELECT 1 FROM terms WHERE source_text=? AND target_lang=? AND domain=?",
                (source_text, target_lang, domain),
            ).fetchone()
        return row is not None

    def get_unknown(
        self, candidates: List[Dict], target_lang: str, domain: str
    ) -> List[Dict]:
        """Return candidates whose (term, target_lang, domain) are NOT in the DB."""
        unknown = []
        for c in candidates:
            term = c.get("term") or c.get("source_text", "")
            if not term:
                continue
            if not self.exists(term, target_lang, domain):
                unknown.append(c)
        return unknown

    def get_top_terms(
        self, target_lang: str, domain: str, top_n: int = 30
    ) -> List[Term]:
        """Return top N injection-safe terms (approved by default; optionally high-confidence unverified)."""
        where = "target_lang=? AND domain=? AND status='approved'"
        params: list = [target_lang, domain]
        if TERM_INJECT_HIGH_CONFIDENCE_UNVERIFIED:
            where = (
                "target_lang=? AND domain=?"
                " AND (status='approved' OR (status='unverified' AND confidence>=?))"
            )
            params = [target_lang, domain, TERM_INJECT_CONF_THRESHOLD]
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT * FROM terms
                WHERE {where}
                ORDER BY usage_count DESC
                LIMIT ?
                """,
                params + [top_n],
            ).fetchall()
        return [_row_to_term(r) for r in rows]

    def get_document_terms(
        self,
        target_lang: str,
        domain: str,
        source_texts: List[str],
    ) -> List[Term]:
        """Return injection-safe terms matching the given source_texts.

        Only returns terms with status='approved' by default.
        When TERM_INJECT_HIGH_CONFIDENCE_UNVERIFIED is true, also includes
        unverified terms with confidence >= TERM_INJECT_CONF_THRESHOLD.
        """
        if not source_texts:
            return []
        placeholders = ",".join("?" for _ in source_texts)
        status_clause = "status='approved'"
        base_params: list = [target_lang, domain]
        if TERM_INJECT_HIGH_CONFIDENCE_UNVERIFIED:
            status_clause = "(status='approved' OR (status='unverified' AND confidence>=?))"
            base_params = [target_lang, domain, TERM_INJECT_CONF_THRESHOLD]
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT * FROM terms
                WHERE target_lang=? AND domain=?
                  AND {status_clause}
                  AND source_text IN ({placeholders})
                ORDER BY confidence DESC, usage_count DESC
                """,
                base_params + list(source_texts),
            ).fetchall()
        return [_row_to_term(r) for r in rows]

    def get_similar_terms_by_embedding(
        self,
        query_vectors: List[List[float]],
        target_lang: str,
        domain: str,
        threshold: float,
        embed_fn: Callable[[List[str]], List[List[float]]],
    ) -> List[Term]:
        """Return DB terms whose source_text is cosine-similar to any query vector.

        Steps:
          1. Load candidate terms (approved + optional high-conf unverified) for
             (target_lang, domain) using the existing injectable-status policy.
          2. Embed each candidate's source_text via the injected embed_fn (no
             vectors are persisted — v1 computes on the fly).
          3. Compute cosine similarity between every (query, candidate) pair.
          4. Return Term objects whose max similarity across all queries is >=
             threshold, sorted descending by that max score.

        Args:
            query_vectors: Pre-computed embedding vectors for the source segments.
                           Empty list → returns [] immediately (no embed call made).
            target_lang:   Target language filter for candidate DB rows.
            domain:        Domain filter for candidate DB rows.
            threshold:     Minimum cosine similarity to include a term.
            embed_fn:      Callable(texts) -> list[list[float]]; called with
                           candidate source_text strings.  Must return [] on failure
                           (caller is responsible for non-fatal behaviour).

        Returns:
            List of Term objects (no duplicates), ordered by descending max score.
        """
        import numpy as np

        if not query_vectors:
            return []

        # Reuse the same status policy as get_document_terms / get_top_terms.
        status_clause = "status='approved'"
        params: list = [target_lang, domain]
        if TERM_INJECT_HIGH_CONFIDENCE_UNVERIFIED:
            status_clause = "(status='approved' OR (status='unverified' AND confidence>=?))"
            params = [target_lang, domain, TERM_INJECT_CONF_THRESHOLD]

        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(
                f"""
                SELECT * FROM terms
                WHERE target_lang=? AND domain=?
                  AND {status_clause}
                ORDER BY confidence DESC, usage_count DESC
                """,
                params,
            ).fetchall()

        if not rows:
            return []

        candidates = [_row_to_term(r) for r in rows]
        candidate_texts = [t.source_text for t in candidates]

        # Embed all candidate source texts via the injected callable.
        candidate_vectors = embed_fn(candidate_texts)
        if not candidate_vectors or len(candidate_vectors) != len(candidates):
            # Embedding failure or mismatched lengths → skip injection.
            return []

        # Convert to numpy arrays for vectorised cosine computation.
        q_mat = np.array(query_vectors, dtype=float)      # shape: (Q, D)
        c_mat = np.array(candidate_vectors, dtype=float)  # shape: (C, D)

        # Normalise rows (add epsilon to avoid division by zero).
        q_norms = np.linalg.norm(q_mat, axis=1, keepdims=True) + 1e-9
        c_norms = np.linalg.norm(c_mat, axis=1, keepdims=True) + 1e-9
        q_norm = q_mat / q_norms
        c_norm = c_mat / c_norms

        # cosine_matrix[q, c] = cosine(query_q, candidate_c)
        cosine_matrix = q_norm @ c_norm.T  # shape: (Q, C)

        # For each candidate, take the max similarity over all queries.
        max_scores = cosine_matrix.max(axis=0)  # shape: (C,)

        # Collect hits above threshold, sorted by descending score.
        hits = [
            (candidates[i], float(max_scores[i]))
            for i in range(len(candidates))
            if float(max_scores[i]) >= threshold
        ]
        hits.sort(key=lambda x: x[1], reverse=True)
        return [term for term, _score in hits]

    # ------------------------------------------------------------------
    # Write methods
    # ------------------------------------------------------------------

    def insert(self, term: Term, strategy: str = "skip") -> str:
        """Insert a term with the given conflict strategy.

        Strategies:
          skip      – never overwrite (safe default)
          overwrite – update unverified records; PROTECTS approved records
          merge     – overwrite only if imported confidence > existing; PROTECTS approved
          force     – overwrites everything, including approved (intentional correction)

        Returns: 'inserted', 'skipped', or 'overwritten'.
        """
        with _DB_LOCK, self._connect() as conn:
            existing = conn.execute(
                "SELECT id, confidence, status FROM terms WHERE source_text=? AND target_lang=? AND domain=?",
                (term.source_text, term.target_lang, term.domain),
            ).fetchone()

            if existing is None:
                conn.execute(
                    """
                    INSERT INTO terms
                        (source_text, target_text, source_lang, target_lang,
                         domain, context_snippet, confidence, usage_count, status, created_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        term.source_text,
                        term.target_text,
                        term.source_lang,
                        term.target_lang,
                        term.domain,
                        term.context_snippet,
                        term.confidence,
                        term.usage_count,
                        term.status,
                        term.created_at or _now_iso(),
                    ),
                )
                conn.commit()
                return "inserted"

            existing_approved = existing["status"] == "approved"
            existing_rejected = existing["status"] == "rejected"

            # Conflict resolution
            if strategy == "skip":
                return "skipped"

            if strategy == "overwrite":
                # Protect approved and rejected records — use 'force' for intentional correction
                if existing_approved or existing_rejected:
                    return "skipped"
                conn.execute(
                    """
                    UPDATE terms SET
                        target_text=?, source_lang=?, context_snippet=?,
                        confidence=?, usage_count=?, status=?, created_at=?
                    WHERE source_text=? AND target_lang=? AND domain=?
                    """,
                    (
                        term.target_text, term.source_lang, term.context_snippet,
                        term.confidence, term.usage_count, term.status,
                        term.created_at or _now_iso(),
                        term.source_text, term.target_lang, term.domain,
                    ),
                )
                conn.commit()
                return "overwritten"

            if strategy == "force":
                # Intentional correction — overwrites approved records too
                conn.execute(
                    """
                    UPDATE terms SET
                        target_text=?, source_lang=?, context_snippet=?,
                        confidence=?, usage_count=?, status=?, created_at=?
                    WHERE source_text=? AND target_lang=? AND domain=?
                    """,
                    (
                        term.target_text, term.source_lang, term.context_snippet,
                        term.confidence, term.usage_count, term.status,
                        term.created_at or _now_iso(),
                        term.source_text, term.target_lang, term.domain,
                    ),
                )
                conn.commit()
                return "overwritten"

            if strategy == "merge":
                # Protect approved and rejected; for unverified/needs_review: take higher confidence
                if existing_approved or existing_rejected:
                    return "skipped"
                existing_confidence = existing["confidence"]
                if term.confidence > existing_confidence:
                    conn.execute(
                        """
                        UPDATE terms SET
                            target_text=?, source_lang=?, context_snippet=?,
                            confidence=?, status=?, created_at=?
                        WHERE source_text=? AND target_lang=? AND domain=?
                        """,
                        (
                            term.target_text, term.source_lang, term.context_snippet,
                            term.confidence, term.status,
                            term.created_at or _now_iso(),
                            term.source_text, term.target_lang, term.domain,
                        ),
                    )
                    conn.commit()
                    return "overwritten"
                return "skipped"

        return "skipped"

    def edit_term(
        self,
        source_text: str,
        target_lang: str,
        domain: str,
        *,
        target_text: str,
        confidence: Optional[float] = None,
    ) -> bool:
        """Update target_text (and optionally confidence) of any term, keeping status='approved'.

        Returns True if the term was found and updated.
        """
        with _DB_LOCK, self._connect() as conn:
            if confidence is not None:
                cur = conn.execute(
                    """
                    UPDATE terms SET target_text=?, confidence=?, status='approved'
                    WHERE source_text=? AND target_lang=? AND domain=?
                    """,
                    (target_text, confidence, source_text, target_lang, domain),
                )
            else:
                cur = conn.execute(
                    """
                    UPDATE terms SET target_text=?, status='approved'
                    WHERE source_text=? AND target_lang=? AND domain=?
                    """,
                    (target_text, source_text, target_lang, domain),
                )
            conn.commit()
        return cur.rowcount > 0

    def approve(self, source_text: str, target_lang: str, domain: str) -> bool:
        """Set status='approved' for a term. Returns True if the term was found."""
        with _DB_LOCK, self._connect() as conn:
            cur = conn.execute(
                "UPDATE terms SET status='approved' WHERE source_text=? AND target_lang=? AND domain=?",
                (source_text, target_lang, domain),
            )
            conn.commit()
        return cur.rowcount > 0

    def reject(self, source_text: str, target_lang: str, domain: str) -> bool:
        """Transition term to rejected status. Returns False if term not found."""
        with _DB_LOCK, self._connect() as conn:
            cur = conn.execute(
                "UPDATE terms SET status='rejected' WHERE source_text=? AND target_lang=? AND domain=?",
                (source_text, target_lang, domain),
            )
            conn.commit()
        return cur.rowcount > 0

    def flag_needs_review(self, source_text: str, target_lang: str, domain: str) -> bool:
        """Transition term to needs_review status. Returns False if term not found."""
        with _DB_LOCK, self._connect() as conn:
            cur = conn.execute(
                "UPDATE terms SET status='needs_review' WHERE source_text=? AND target_lang=? AND domain=?",
                (source_text, target_lang, domain),
            )
            conn.commit()
        return cur.rowcount > 0

    def get_unverified(
        self,
        target_lang: Optional[str] = None,
        domain: Optional[str] = None,
    ) -> List[Term]:
        """Return all unverified terms, optionally filtered."""
        conditions = ["status='unverified'"]
        params: list = []
        if target_lang:
            conditions.append("target_lang=?")
            params.append(target_lang)
        if domain:
            conditions.append("domain=?")
            params.append(domain)
        sql = f"SELECT * FROM terms WHERE {' AND '.join(conditions)} ORDER BY id DESC"
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [_row_to_term(r) for r in rows]

    def increment_usage(self, source_text: str, target_lang: str, domain: str) -> None:
        """Increment usage_count for a term."""
        with _DB_LOCK, self._connect() as conn:
            conn.execute(
                """
                UPDATE terms SET usage_count = usage_count + 1
                WHERE source_text=? AND target_lang=? AND domain=?
                """,
                (source_text, target_lang, domain),
            )
            conn.commit()

    # ------------------------------------------------------------------
    # Statistics
    # ------------------------------------------------------------------

    def get_stats(self) -> Dict:
        """Return total / unverified / by_target_lang / by_domain / by_status statistics."""
        with _DB_LOCK, self._connect() as conn:
            total = conn.execute("SELECT COUNT(*) FROM terms").fetchone()[0]
            unverified = conn.execute(
                "SELECT COUNT(*) FROM terms WHERE status='unverified'"
            ).fetchone()[0]
            by_lang_rows = conn.execute(
                "SELECT target_lang, COUNT(*) as cnt FROM terms GROUP BY target_lang"
            ).fetchall()
            by_domain_rows = conn.execute(
                "SELECT domain, COUNT(*) as cnt FROM terms GROUP BY domain"
            ).fetchall()
            status_rows = conn.execute(
                "SELECT status, COUNT(*) FROM terms GROUP BY status"
            ).fetchall()
        by_status = {row[0]: row[1] for row in status_rows}
        return {
            "total": total,
            "unverified": unverified,
            "by_target_lang": {r["target_lang"]: r["cnt"] for r in by_lang_rows},
            "by_domain": {r["domain"]: r["cnt"] for r in by_domain_rows},
            "by_status": by_status,
        }

    # ------------------------------------------------------------------
    # Export
    # ------------------------------------------------------------------

    def get_approved(
        self,
        target_lang: Optional[str] = None,
        domain: Optional[str] = None,
    ) -> List[Term]:
        """Return all approved terms, optionally filtered."""
        conditions = ["status='approved'"]
        params: list = []
        if target_lang:
            conditions.append("target_lang=?")
            params.append(target_lang)
        if domain:
            conditions.append("domain=?")
            params.append(domain)
        sql = f"SELECT * FROM terms WHERE {' AND '.join(conditions)} ORDER BY id DESC"
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [_row_to_term(r) for r in rows]

    def get_rejected(
        self,
        target_lang: Optional[str] = None,
        domain: Optional[str] = None,
    ) -> List[Term]:
        """Return all rejected terms, optionally filtered."""
        conditions = ["status='rejected'"]
        params: list = []
        if target_lang:
            conditions.append("target_lang=?")
            params.append(target_lang)
        if domain:
            conditions.append("domain=?")
            params.append(domain)
        sql = f"SELECT * FROM terms WHERE {' AND '.join(conditions)} ORDER BY id DESC"
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [_row_to_term(r) for r in rows]

    def export_json(self, path: Path, status_filter: Optional[str] = None) -> None:
        """Export terms to a JSON file. status_filter: 'approved' | 'unverified' | None (all)."""
        terms = self._all_terms(status_filter)
        payload = {
            "version": 1,
            "exported_at": _now_iso(),
            "terms": [_term_to_dict(t) for t in terms],
        }
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        with path.open("w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)

    def export_csv(self, path: Path, status_filter: Optional[str] = None) -> None:
        """Export terms to a CSV file. status_filter: 'approved' | 'unverified' | None (all)."""
        terms = self._all_terms(status_filter)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)
        fieldnames = [
            "source_text", "target_text", "source_lang", "target_lang",
            "domain", "context_snippet", "confidence", "usage_count", "status",
        ]
        with path.open("w", encoding="utf-8", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            for t in terms:
                writer.writerow({k: getattr(t, k) for k in fieldnames})

    def export_xlsx(self, path: Path, status_filter: Optional[str] = None) -> None:
        """Export terms to an XLSX file. status_filter: 'approved' | 'unverified' | None (all)."""
        try:
            import openpyxl
        except ImportError:
            raise RuntimeError("openpyxl is required for XLSX export. Install it with: pip install openpyxl")

        terms = self._all_terms(status_filter)
        path = Path(path)
        path.parent.mkdir(parents=True, exist_ok=True)

        # Group by target_lang
        by_lang: Dict[str, List[Term]] = {}
        for t in terms:
            by_lang.setdefault(t.target_lang, []).append(t)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # Remove default sheet

        headers = [
            "source_text", "target_text", "source_lang", "target_lang",
            "domain", "context_snippet", "confidence", "usage_count", "status",
        ]
        for lang, lang_terms in by_lang.items():
            ws = wb.create_sheet(title=lang[:31])  # Sheet names max 31 chars
            ws.append(headers)
            for t in lang_terms:
                ws.append([getattr(t, h) for h in headers])

        if not wb.worksheets:
            wb.create_sheet("terms")

        wb.save(str(path))

    # ------------------------------------------------------------------
    # Import
    # ------------------------------------------------------------------

    def import_file(self, path: Path, strategy: str = "skip") -> Dict[str, int]:
        """Import terms from a JSON or CSV file.

        Returns counts: inserted / skipped / overwritten.
        """
        path = Path(path)
        suffix = path.suffix.lower()

        if suffix == ".json":
            terms = self._parse_json_import(path)
        elif suffix == ".csv":
            terms = self._parse_csv_import(path)
        else:
            raise ValueError(f"Unsupported import format: {suffix}")

        counts: Dict[str, int] = {"inserted": 0, "skipped": 0, "overwritten": 0}
        for term in terms:
            result = self.insert(term, strategy=strategy)
            counts[result] = counts.get(result, 0) + 1

        logger.info(
            "[TermDB] Import complete: inserted=%d skipped=%d overwritten=%d",
            counts["inserted"], counts["skipped"], counts["overwritten"],
        )
        return counts

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    def _all_terms(self, status_filter: Optional[str] = None) -> List[Term]:
        sql = "SELECT * FROM terms"
        params: list = []
        if status_filter in _VALID_STATUSES:
            sql += " WHERE status=?"
            params.append(status_filter)
        sql += " ORDER BY id"
        with _DB_LOCK, self._connect() as conn:
            rows = conn.execute(sql, params).fetchall()
        return [_row_to_term(r) for r in rows]

    def _parse_json_import(self, path: Path) -> List[Term]:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f)
        raw_terms = data.get("terms", data) if isinstance(data, dict) else data
        return [_dict_to_term(d) for d in raw_terms if isinstance(d, dict)]

    def _parse_csv_import(self, path: Path) -> List[Term]:
        terms: List[Term] = []
        with path.open("r", encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                terms.append(_dict_to_term(dict(row)))
        return terms


# ------------------------------------------------------------------
# Helpers
# ------------------------------------------------------------------

def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _row_to_term(row: sqlite3.Row) -> Term:
    return Term(
        source_text=row["source_text"],
        target_text=row["target_text"],
        source_lang=row["source_lang"],
        target_lang=row["target_lang"],
        domain=row["domain"],
        context_snippet=row["context_snippet"],
        confidence=float(row["confidence"]),
        usage_count=int(row["usage_count"]),
        status=row["status"] if "status" in row.keys() else "unverified",
        created_at=row["created_at"],
    )


def _term_to_dict(t: Term) -> Dict:
    return {
        "source_text": t.source_text,
        "target_text": t.target_text,
        "source_lang": t.source_lang,
        "target_lang": t.target_lang,
        "domain": t.domain,
        "context_snippet": t.context_snippet,
        "confidence": t.confidence,
        "usage_count": t.usage_count,
        "status": t.status,
    }


def _dict_to_term(d: Dict) -> Term:
    """Convert an import dict to a Term. Defaults status to 'approved' (human-curated import)."""
    return Term(
        source_text=str(d.get("source_text", "")),
        target_text=str(d.get("target_text", "")),
        source_lang=str(d.get("source_lang", "")),
        target_lang=str(d.get("target_lang", "")),
        domain=str(d.get("domain", "general")),
        context_snippet=str(d.get("context_snippet", "")),
        confidence=float(d.get("confidence", 1.0)),
        usage_count=int(d.get("usage_count", 0)),
        status=str(d.get("status", "approved")),  # imports are human-curated
        created_at=str(d.get("created_at") or _now_iso()),
    )
