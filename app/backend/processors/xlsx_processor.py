"""XLSX/XLS translation processor."""

from __future__ import annotations

import os
import threading
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment

from app.backend.clients.ollama_client import OllamaClient
from app.backend.config import DEFAULT_MAX_BATCH_CHARS, EXCEL_FORMULA_MODE, MAX_SEGMENTS, MAX_TEXT_LENGTH
from app.backend.processors.com_helpers import excel_convert, is_win32com_available
from app.backend.processors.libreoffice_helpers import is_libreoffice_available, xls_to_xlsx
from app.backend.services.translation_service import translate_texts
from app.backend.utils.exceptions import check_document_size_limits
from app.backend.utils.logging_utils import logger
from app.backend.utils.text_utils import normalize_text, should_translate


def _get_display_text_for_translation(ws: Any, ws_vals: Optional[Any], r: int, c: int) -> Optional[str]:
    val = ws.cell(row=r, column=c).value
    if isinstance(val, str) and val.startswith("="):
        if ws_vals is not None:
            shown = ws_vals.cell(row=r, column=c).value
            return shown if isinstance(shown, str) and shown.strip() else None
        return None
    if isinstance(val, str) and val.strip():
        return val
    if ws_vals is not None:
        shown = ws_vals.cell(row=r, column=c).value
        if isinstance(shown, str) and shown.strip():
            return shown
    return None


def translate_xlsx_xls(
    in_path: str,
    out_path: str,
    targets: List[str],
    src_lang: Optional[str],
    client: OllamaClient,
    excel_formula_mode: str = EXCEL_FORMULA_MODE,
    stop_flag: Optional[threading.Event] = None,
    log: Callable[[str], None] = lambda s: None,
    max_segments: int = MAX_SEGMENTS,
    max_text_length: int = MAX_TEXT_LENGTH,
    max_batch_chars: int = DEFAULT_MAX_BATCH_CHARS,
    pre_translate_hook: Optional[Callable[[List[str]], None]] = None,
    post_translate_hook: Optional[Callable[[List[Tuple[str, str, str]]], None]] = None,
    terms_getter: Optional[Callable[[], list]] = None,
    block_overrides: Optional[Dict[str, str]] = None,
    status_callback: Optional[Callable[[Optional[str]], None]] = None,
) -> bool:
    ext = Path(in_path).suffix.lower()
    out_xlsx = Path(out_path).with_suffix(".xlsx")
    if ext == ".xls":
        tmp = str(Path(out_path).with_suffix("")) + "__from_xls.xlsx"
        if is_libreoffice_available():
            log("[XLS] Converting to .xlsx via LibreOffice")
            xls_to_xlsx(in_path, tmp)
        elif is_win32com_available():
            log("[XLS] Converting to .xlsx via COM")
            excel_convert(in_path, tmp)
        else:
            raise RuntimeError(
                "Cannot convert .xls: neither LibreOffice nor Excel COM is "
                "available. Install LibreOffice: "
                "sudo apt install libreoffice-core (Linux) / "
                "brew install --cask libreoffice (macOS)"
            )
        try:
            return translate_xlsx_xls(
                tmp,
                out_path,
                targets,
                src_lang,
                client,
                excel_formula_mode=excel_formula_mode,
                stop_flag=stop_flag,
                log=log,
                max_segments=max_segments,
                max_text_length=max_text_length,
                max_batch_chars=max_batch_chars,
                pre_translate_hook=pre_translate_hook,
                post_translate_hook=post_translate_hook,
                block_overrides=block_overrides,
            )
        finally:
            try:
                os.remove(tmp)
            except OSError as exc:
                logger.debug("Failed to remove temp file %s: %s", tmp, exc)
    if ext != ".xlsx":
        raise RuntimeError("Unsupported Excel type")

    wb = openpyxl.load_workbook(in_path, data_only=False)
    try:
        wb_vals = openpyxl.load_workbook(in_path, data_only=True)
    except (OSError, ValueError, KeyError) as exc:
        logger.debug("Failed to load workbook with data_only=True: %s", exc)
        wb_vals = None

    segs: List[Tuple[str, int, int, str, bool]] = []
    total_text_length = 0
    for ws in wb.worksheets:
        ws_vals = wb_vals[ws.title] if wb_vals and ws.title in wb_vals.sheetnames else None
        max_row, max_col = ws.max_row, ws.max_column
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                src_text = _get_display_text_for_translation(ws, ws_vals, r, c)
                if not src_text:
                    continue
                if not should_translate(src_text, (src_lang or "auto")):
                    continue
                val = ws.cell(row=r, column=c).value
                is_formula = isinstance(val, str) and val.startswith("=")
                segs.append((ws.title, r, c, src_text, is_formula))
                total_text_length += len(src_text)

    check_document_size_limits(
        segment_count=len(segs),
        total_text_length=total_text_length,
        max_segments=max_segments,
        max_text_length=max_text_length,
        document_type="Excel document",
    )

    log(f"[Excel] cells: {len(segs)}")
    # Preserve document order for better batch context (used for pre/post hooks)
    _seen: set[str] = set()
    uniq: list[str] = []
    for s in segs:
        if s[3] not in _seen:
            _seen.add(s[3])
            uniq.append(s[3])
    if pre_translate_hook:
        pre_translate_hook(uniq)
    _terms = terms_getter() if terms_getter else None

    # p3-llm-judge: block_overrides seam
    _ext_name = os.path.splitext(in_path)[1].lstrip(".")
    _file_stem = os.path.splitext(os.path.basename(in_path))[0]
    stopped = False

    # IP-4 (BR-81): tmap uses 3-element key (tgt, src_text, col) where col is 0-based.
    # Each worksheet is treated as a table; one translate_once call per worksheet per target.
    tmap: Dict = {}

    if block_overrides is not None:
        # For block_overrides, map each block's text to all cells with matching text.
        idx_map: Dict[str, str] = {}
        for idx, src_text in enumerate(uniq):
            block_id = f"{_ext_name}:{_file_stem}:{idx}"
            if block_id in block_overrides:
                idx_map[src_text] = block_overrides[block_id]
        for sheet_name, r, c, src_text, is_formula in segs:
            c0 = c - 1  # 0-based
            for tgt in targets:
                tmap[(tgt, src_text, c0)] = idx_map.get(src_text, src_text)
        log(f"[Excel] block_overrides applied: {len(block_overrides)} overrides, {len(uniq)} blocks")
    else:
        # IP-4: per-worksheet serialization — one translate_once call per worksheet per target.
        from collections import defaultdict
        from dataclasses import dataclass as _dc
        from app.backend.utils import table_serializer

        @_dc
        class _XlCellProxy:
            row: int
            col: int
            content: str
            is_numeric: bool = False

        # Group segs by worksheet
        sheet_segs: Dict = defaultdict(list)
        for seg in segs:
            sheet_segs[seg[0]].append(seg)

        for sheet_name, s_segs in sheet_segs.items():
            if stop_flag and stop_flag.is_set():
                stopped = True
                break
            ws = wb[sheet_name]
            num_rows = ws.max_row
            num_cols = ws.max_column

            # Map (r-1, c-1) → text for non-formula cells
            cells_by_pos = {(seg[1] - 1, seg[2] - 1): seg[3] for seg in s_segs if not seg[4]}

            proxy_cells = [
                _XlCellProxy(row=r0, col=c0, content=cells_by_pos.get((r0, c0), ""))
                for r0 in range(num_rows) for c0 in range(num_cols)
            ]

            for tgt in targets:
                if stop_flag and stop_flag.is_set():
                    stopped = True
                    break
                src_for_prompt = src_lang or "auto"
                serialized = table_serializer.serialize(proxy_cells)
                prompt = client._build_table_translate_prompt(serialized, src_for_prompt, tgt)
                # grid stays None on any error → fallback always runs (BR-82)
                grid = None
                try:
                    ok, response = client.translate_once(prompt, tgt, src_lang)
                    if ok:
                        grid = table_serializer.parse(response, num_rows, num_cols)
                        if grid is None:
                            logger.warning(
                                "[Excel] Sheet '%s': parse() returned None (expected %d×%d); "
                                "falling back to per-cell batch for target=%s",
                                sheet_name, num_rows, num_cols, tgt,
                            )
                except Exception as exc:
                    logger.warning(
                        "[Excel] Sheet '%s' translate_once failed (target=%s): %s; "
                        "falling back to per-cell batch",
                        sheet_name, tgt, exc,
                    )
                if grid is not None:
                    for ws_name, r, c, src_text, is_formula in s_segs:
                        if not is_formula:
                            r0, c0 = r - 1, c - 1
                            tmap[(tgt, src_text, c0)] = grid[r0][c0]
                else:
                    # Fallback: per-cell batch (BR-82)
                    fallback_texts = list(dict.fromkeys(
                        s[3] for s in s_segs
                        if not s[4] and should_translate(s[3], src_for_prompt)
                    ))
                    if fallback_texts:
                        fb_tmap, _, _, _ = translate_texts(
                            fallback_texts, [tgt], src_lang, client,
                            max_batch_chars=max_batch_chars,
                            stop_flag=stop_flag, log=log,
                        )
                        for (fb_tgt, fb_text), fb_tr in fb_tmap.items():
                            for ws_name, r, c, src_text, is_formula in s_segs:
                                if src_text == fb_text:
                                    tmap[(fb_tgt, fb_text, c - 1)] = fb_tr

    for sheet_name, r, c, src_text, is_formula in segs:
        c0 = c - 1  # 0-based column index (BR-81 dedup key)
        if not all((tgt, src_text, c0) in tmap for tgt in targets):
            continue
        ws = wb[sheet_name]
        trs = [tmap[(tgt, src_text, c0)] for tgt in targets]
        if is_formula:
            if excel_formula_mode == "skip":
                continue
            if excel_formula_mode == "comment":
                txt_comment = "\n".join([f"[{t}] {res}" for t, res in zip(targets, trs)])
                cell = ws.cell(row=r, column=c)
                exist = cell.comment
                if not exist or normalize_text(exist.text) != normalize_text(txt_comment):
                    cell.comment = Comment(txt_comment, "translator")
                continue
            continue
        combined = "\n".join([src_text] + trs)
        cell = ws.cell(row=r, column=c)
        if isinstance(cell.value, str) and normalize_text(cell.value) == normalize_text(combined):
            continue
        cell.value = combined
        try:
            if cell.alignment:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical=cell.alignment.vertical,
                    wrap_text=True,
                )
            else:
                cell.alignment = Alignment(wrap_text=True)
        except (TypeError, AttributeError) as exc:
            logger.debug("Cell alignment copy failed: %s", exc)
            cell.alignment = Alignment(wrap_text=True)

    wb.save(out_xlsx)

    if post_translate_hook is not None:
        # Build a flat (tgt, src_text) → translation lookup from the 3-element tmap
        # by taking the first matching cell translation for each unique text.
        flat_tmap: Dict = {}
        for (tgt, src_text, col), tr in tmap.items():
            if (tgt, src_text) not in flat_tmap:
                flat_tmap[(tgt, src_text)] = tr
        tuples: List[Tuple[str, str, str]] = []
        for idx, src_text in enumerate(uniq):
            for tgt in targets:
                if (tgt, src_text) in flat_tmap:
                    tuples.append((f"{_ext_name}:{_file_stem}:{idx}", src_text, flat_tmap[(tgt, src_text)]))
        if tuples:
            post_translate_hook(tuples)

    if stopped:
        log(f"[Excel] partial output: {out_xlsx.name}")
    else:
        log(f"[Excel] output: {out_xlsx.name}")

    return stopped
